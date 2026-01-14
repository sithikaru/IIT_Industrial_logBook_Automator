import streamlit as st
import pandas as pd
import os
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Alignment
from io import BytesIO
import time
from groq import Groq

# --- CONFIGURATION & DATA ---
FILE_NAME = "my_placement_logs.csv"
# Expanded list of activities based on your document
ACTIVITIES = {
    "1.1": "Conduct preliminary investigations",
    "1.2": "Carry out feasibility study",
    "2.1": "Analyze current system",
    "2.2": "Identify requirements",
    "3.1": "Design data (ERD, DFD)",
    "3.2": "Design process outlines",
    "4.1": "Program design",
    "4.2": "Program code",
    "4.3": "Test programs",
    "5.1": "Testing module",
    "5.2": "Integration testing",
    "6.1": "Educate and train users",
    "9.1": "Maintenance/Bug Fixing",
    "12.1": "Project Management",
    "19.1": "Cybersecurity / Security",
    "22.1": "Cloud Computing Tasks",
    "Other": "General / Administrative"
}

# --- HELPER FUNCTIONS ---
def load_data():
    if not os.path.exists(FILE_NAME):
        df = pd.DataFrame(columns=["Date", "Day", "Week_Ending", "Activity_Code", "Description", "Problems", "Solutions"])
        df.to_csv(FILE_NAME, index=False)
        return df
    return pd.read_csv(FILE_NAME)

def save_entry(date_obj, activity_code, desc, prob, sol):
    df = load_data()
    day_name = date_obj.strftime("%A")
    # Logic: Week ends on the upcoming Sunday
    days_ahead = 6 - date_obj.weekday()
    week_ending = date_obj + timedelta(days=days_ahead)
    
    new_entry = {
        "Date": date_obj,
        "Day": day_name.upper(),
        "Week_Ending": week_ending.strftime("%Y-%m-%d"),
        "Activity_Code": activity_code,
        "Description": desc,
        "Problems": prob if prob else "",
        "Solutions": sol if sol else ""
    }
    df = pd.concat([df, pd.DataFrame([new_entry])], ignore_index=True)
    df.to_csv(FILE_NAME, index=False)
    df.to_csv(FILE_NAME, index=False)
    # st.success("Entry Saved!") # Removed to avoid clutter in bulk mode, handle in UI
    
def get_writeable_cell(ws, row, col):
    """
    Returns the writeable cell (top-left) if the target is a merged cell.
    """
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, openpyxl.cell.cell.MergedCell):
        for merged_range in ws.merged_cells.ranges:
            if (col >= merged_range.min_col and col <= merged_range.max_col and
                row >= merged_range.min_row and row <= merged_range.max_row):
                return ws.cell(row=merged_range.min_row, column=merged_range.min_col)
    return cell

def get_week_start(date_obj):
    """Returns the Monday of the week for the given date."""
    return date_obj - timedelta(days=date_obj.weekday())

import copy

def copy_range(ws, src_min_row, src_max_row, src_min_col, src_max_col, dest_min_row):
    """
    Copies a range of cells (values + styles + merges) to a new row offset.
    Returns the number of rows copied.
    """
    rows_count = src_max_row - src_min_row + 1
    dest_max_row = dest_min_row + rows_count - 1
    dest_min_col = src_min_col
    dest_max_col = src_max_col
    
    # 0. Cleanup Destination Merges
    # If the destination has existing merges, we must unmerge them first to allow writing values.
    # Otherwise we hit 'MergedCell' read-only errors.
    for merged_range in list(ws.merged_cells.ranges):
        # Check for overlap
        if (merged_range.min_row <= dest_max_row and merged_range.max_row >= dest_min_row and
            merged_range.min_col <= dest_max_col and merged_range.max_col >= dest_min_col):
            try:
                ws.unmerge_cells(start_row=merged_range.min_row, start_column=merged_range.min_col,
                                 end_row=merged_range.max_row, end_column=merged_range.max_col)
            except KeyError:
                # Cell might be missing from internal index if rows were deleted beforehand
                pass

    # 1. Copy Cells
    for row_offset in range(rows_count):
        src_row = src_min_row + row_offset
        dest_row = dest_min_row + row_offset
        
        for col in range(src_min_col, src_max_col + 1):
            src_cell = ws.cell(row=src_row, column=col)
            dest_cell = ws.cell(row=dest_row, column=col)
            
            # Copy value
            dest_cell.value = src_cell.value
            
            # Copy style (simplified: alignment, font, border, fill)
            if src_cell.has_style:
                dest_cell.font = copy.copy(src_cell.font)
                dest_cell.border = copy.copy(src_cell.border)
                dest_cell.fill = copy.copy(src_cell.fill)
                dest_cell.number_format = copy.copy(src_cell.number_format)
                dest_cell.protection = copy.copy(src_cell.protection)
                dest_cell.alignment = copy.copy(src_cell.alignment)

    # 2. Copy Merged Cells
    # We need to find merges in the source range and map them to the dest range
    # Iterate over a COPY of the ranges because merge_cells modifies the collection
    for merged_range in list(ws.merged_cells.ranges):
        if (merged_range.min_row >= src_min_row and 
            merged_range.max_row <= src_max_row and
            merged_range.min_col >= src_min_col and 
            merged_range.max_col <= src_max_col):
            
            # Calculate offset
            offset_row = dest_min_row - src_min_row
            
            new_min_row = merged_range.min_row + offset_row
            new_max_row = merged_range.max_row + offset_row
            new_min_col = merged_range.min_col
            new_max_col = merged_range.max_col
            
            ws.merge_cells(start_row=new_min_row, start_column=new_min_col, 
                           end_row=new_max_row, end_column=new_max_col)
            
    return rows_count

def fill_excel_sheet(template_file, data_df, start_date, end_date, output_path=None):
    """
    Refactored to:
    1. Create one sheet per Month between start_date and end_date.
    2. Dynamically generate 4 or 5 tables per sheet based on Sundays.
    3. Fill tables with data for that month.
    """
    wb = openpyxl.load_workbook(template_file)
    
    # Identify Template Sheet
    if 'Logs' in wb.sheetnames:
        template_ws = wb['Logs']
    else:
        template_ws = wb.active
        
    template_ws.title = "Template" # Rename for clarity
    
    # --- 1. Identify Template Range ---
    start_row = None
    for row in range(1, 100):
        c = template_ws.cell(row=row, column=1)
        if c.value and "WEEK ENDING" in str(c.value).upper():
            start_row = row
            break
            
    if not start_row:
        return None, "Could not find 'WEEK ENDING' in the template."

    TEMPLATE_ROW_COUNT = 21 # Assumed block size
    
    # Convert dates
    data_df['Week_Ending_Dt'] = pd.to_datetime(data_df['Week_Ending'])
    
    current_date = start_date.replace(day=1)
    
    # Iterate Months
    while current_date <= end_date:
        month_name = current_date.strftime("%b %Y")
        
        # Create new sheet from template
        new_ws = wb.copy_worksheet(template_ws)
        new_ws.title = month_name
        
        # --- CLEANUP: Keep only the first template block ---
        # We assume the first block (start_row to +TEMPLATE_ROW_COUNT) is the master.
        # Delete everything below it to avoid junk from the template file.
        cutoff_row = start_row + TEMPLATE_ROW_COUNT
        rows_to_delete = new_ws.max_row - cutoff_row + 10
        if rows_to_delete > 0:
            new_ws.delete_rows(cutoff_row, amount=rows_to_delete)
        
        # Get Sundays in this month
        # Start from 1st of month
        curr_mon = current_date
        next_mon = (curr_mon.replace(day=28) + timedelta(days=4)).replace(day=1) # Advance to next month 1st
        
        # Find first Sunday
        d = curr_mon
        while d.weekday() != 6: # 6 = Sunday
            d += timedelta(days=1)
            
        month_sundays = []
        while d < next_mon:
            month_sundays.append(d)
            d += timedelta(days=7)
            
        # Target Sundays is ALL of them (4 or 5)
        target_sundays = month_sundays
        num_tables = len(target_sundays)
        
        # Determine table positions
        tables_start_rows = []
        for i in range(num_tables):
            tables_start_rows.append(start_row + (TEMPLATE_ROW_COUNT + 1) * i)
        
        # Copy template to additional positions
        # Note: Position 0 is already there (from the sheet copy).
        # We copy for i=1 to N-1
        for t_row in tables_start_rows[1:]:
            copy_range(new_ws, start_row, start_row + TEMPLATE_ROW_COUNT - 1, 1, 20, t_row)
        
        # Fill Tables
        day_map = {
            "MONDAY": 2, "TUESDAY": 3, "WEDNESDAY": 4, 
            "THURSDAY": 5, "FRIDAY": 6, "SATURDAY": 7, "SUNDAY": 8
        }
        
        for idx, t_row in enumerate(tables_start_rows):
            # Clear critical cells first (Date, Desc, Code, Probs) - in case Template had junk
            # Date
            c = get_writeable_cell(new_ws, t_row, 2)
            if c: c.value = None
            
            # Prob/Sol
            c = get_writeable_cell(new_ws, t_row + 10, 2)
            if c: c.value = None
            c = get_writeable_cell(new_ws, t_row + 10, 3)
            if c: c.value = None
            
            # Days
            for i in range(2, 9):
                c = get_writeable_cell(new_ws, t_row + i, 2)
                if c: c.value = None
                c = get_writeable_cell(new_ws, t_row + i, 3)
                if c: c.value = None

            if idx < len(target_sundays):
                week_dt = target_sundays[idx]
                week_str = week_dt.strftime("%Y-%m-%d")
                
                # Fill Date
                date_cell = get_writeable_cell(new_ws, t_row, 2)
                if date_cell:
                    date_cell.value = week_str
                    date_cell.alignment = Alignment(horizontal='left')
                
                # Fill Data
                week_data = data_df[data_df['Week_Ending'] == week_str]
                if not week_data.empty:
                    problems_list = []
                    solutions_list = []
                    for _, row_data in week_data.iterrows():
                        day_offset = day_map.get(row_data['Day'], None)
                        if day_offset:
                            cell_desc = get_writeable_cell(new_ws, t_row + day_offset, 2)
                            if cell_desc:
                                cell_desc.value = row_data['Description']
                                cell_desc.alignment = Alignment(wrap_text=True, vertical='top')

                            cell_code = get_writeable_cell(new_ws, t_row + day_offset, 3)
                            if cell_code:
                                cell_code.value = row_data['Activity_Code']
                                cell_code.alignment = Alignment(horizontal='center', vertical='top')
                            
                            if pd.notna(row_data['Problems']) and str(row_data['Problems']).strip():
                                problems_list.append(str(row_data['Problems']))
                            if pd.notna(row_data['Solutions']) and str(row_data['Solutions']).strip():
                                solutions_list.append(str(row_data['Solutions']))

                    if problems_list:
                        cell = get_writeable_cell(new_ws, t_row + 10, 2)
                        if cell:
                            cell.value = "\n".join(problems_list)
                            cell.alignment = Alignment(wrap_text=True, vertical='top')
                    if solutions_list:
                        cell = get_writeable_cell(new_ws, t_row + 10, 3)
                        if cell:
                            cell.value = "\n".join(solutions_list)
                            cell.alignment = Alignment(wrap_text=True, vertical='top')

        # Advance to next month
        current_date = next_mon

    # Move Template to end or hide it?
    # Let's just delete it to be clean, as requested "created ... tabs"
    if 'Template' in wb.sheetnames:
        del wb['Template']

    # Save
    if output_path:
        wb.save(output_path)
        return None, "Saved directly to file."
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output, "Success"

# --- GUI LAYOUT ---
st.set_page_config(page_title="Placement Log Automator", page_icon="🚀", layout="wide")

st.title("🚀 Industrial Placement Log Book Automator")
st.markdown("### Log daily. Generate Excel weekly.")

# Data Loading
df = load_data()

# --- TABS ---
tab_git, tab_daily, tab_manual, tab_excel, tab_hist = st.tabs(["🚀 Bulk Auto-Fill (Git)", "📝 Daily Log", "📚 Manual Weekly Fill", "🤖 Excel Automator", "📊 History"])

import subprocess
import requests
import json
from dotenv import load_dotenv

# Load environment variables
load_dotenv()

# --- TAB 1: GITHUB IMPORT (MAIN) ---
with tab_git:
    st.header("☁️ GitHub History Importer")

    # Load Config from Env
    gh_username = os.getenv("GITHUB_USERNAME", "")
    gh_token = os.getenv("GITHUB_TOKEN", "")
    groq_api_key = os.getenv("GROQ_API_KEY", "")

    if not gh_token:
        st.warning("⚠️ `GITHUB_TOKEN` is missing in `.env` file.")
    if not gh_username:
        st.warning("⚠️ `GITHUB_USERNAME` is missing in `.env` file.")
    if not groq_api_key:
        st.info("💡 `GROQ_API_KEY` is missing. AI summarization will be skipped. Get one free at https://console.groq.com/")

    # 2. Repo Selection
    st.subheader("Select Repositories")
    
    if "my_github_repos" not in st.session_state:
        st.session_state.my_github_repos = []

    col_btn, col_manual = st.columns([1, 2])
    with col_btn:
        if st.button("🔄 Fetch Your Repositories"):
            try:
                found_repos = []
                page = 1
                while True:
                    if gh_token:
                        # Authenticated: Get all accessible repos (private & public)
                        url = "https://api.github.com/user/repos"
                        headers = {"Authorization": f"token {gh_token}", "Accept": "application/vnd.github.v3+json"}
                        params = {"per_page": 100, "page": page, "affiliation": "owner,collaborator,organization_member", "sort": "updated"}
                    else:
                        # Public only
                        url = f"https://api.github.com/users/{gh_username}/repos"
                        headers = {"Accept": "application/vnd.github.v3+json"}
                        params = {"per_page": 100, "page": page, "sort": "updated"}
                    
                    resp = requests.get(url, headers=headers, params=params)
                    if resp.status_code == 200:
                        data = resp.json()
                        if not data:
                            break # No more pages
                        for r in data:
                            found_repos.append(r["full_name"])
                        page += 1
                        # Safety break for massive accounts (limit to 300 for now)
                        if page > 3: 
                            break
                    else:
                        st.error(f"Error fetching repos: {resp.status_code} - {resp.text}")
                        break
                
                if found_repos:
                    st.session_state.my_github_repos = sorted(list(set(found_repos)))
                    st.success(f"Found {len(found_repos)} repositories!")
                else:
                    st.warning("No repositories found.")
                    
            except Exception as e:
                st.error(f"Failed to fetch repositories: {e}")

    # Allow selection from fetched list OR manual entry
    default_options = st.session_state.my_github_repos if st.session_state.my_github_repos else ["owner/repository-name"]
    
    selected_repos = st.multiselect(
        "Choose Repositories", 
        st.session_state.my_github_repos if st.session_state.my_github_repos else default_options,
        default=[] # Start empty so user can choose
    )
    
    # Fallback for manual addition if fetch fails or repo missing
    if not st.session_state.my_github_repos:
        st.caption("Or type manually below if fetch fails:")
        manual_entry = st.text_area("Manual Repo List (owner/repo)", height=68, 
                                    value="owner/repository-name")
        if manual_entry:
            manual_list = [r.strip() for r in manual_entry.split('\n') if r.strip()]
            # Combine unique
            selected_repos = list(set(selected_repos + manual_list))

    # 3. Import Logic
    c_d1, c_d2 = st.columns(2)
    with c_d1:
        start_date = st.date_input("Start Date", datetime.today() - timedelta(days=30))
    with c_d2:
        end_date = st.date_input("End Date", datetime.today())

    scan_all_branches = st.checkbox("🕵️‍♀️ Scan ALL branches (Slow)", value=False, help="Check this to find commits on unmerged feature branches.")
    
    use_author_filter = True
    if gh_username:
        use_author_filter = st.checkbox(f"Filter by author: {gh_username}", value=True, help="Uncheck to see commits from everyone.")

    if st.button("🚀 Fetch & Generate Logs"):
        if not selected_repos:
            st.error("Please select at least one repository.")
        else:
            headers = {"Accept": "application/vnd.github.v3+json"}
            if gh_token:
                headers["Authorization"] = f"token {gh_token}"
            
            all_commits = []
            seen_shas = set() # To store unique commit SHAs

            progress_bar = st.progress(0)
            status_text = st.empty()
            
            total_repos = len(selected_repos)
            
            for i, repo in enumerate(selected_repos):
                # Determine branches to scan
                branches = []
                if scan_all_branches:
                    status_text.text(f"Listing branches for {repo}...")
                    try:
                        br_url = f"https://api.github.com/repos/{repo}/branches"
                        br_resp = requests.get(br_url, headers=headers)
                        if br_resp.status_code == 200:
                            branches = [b["name"] for b in br_resp.json()]
                        else:
                            st.warning(f"Could not list branches for {repo}, defaulting to main.")
                            branches = [None]
                    except:
                        branches = [None]
                else:
                    branches = [None] # None means default branch

                for branch_name in branches:
                    branch_label = branch_name if branch_name else "default"
                    status_text.text(f"Fetching {repo} [{branch_label}]...")
                    
                    try:
                        # Pagination Loop
                        page = 1
                        while True:
                            url = f"https://api.github.com/repos/{repo}/commits"
                            params = {
                                "since": start_date.strftime('%Y-%m-%dT00:00:00Z'),
                                "until": (end_date + timedelta(days=1)).strftime('%Y-%m-%dT00:00:00Z'),
                                "per_page": 100,
                                "page": page
                            }
                            # Apply Author Filter IF checkbox is checked
                            if gh_username and use_author_filter:
                                params["author"] = gh_username
                                
                            if branch_name:
                                params["sha"] = branch_name
                            
                            resp = requests.get(url, headers=headers, params=params)
                            
                            if resp.status_code == 200:
                                commits = resp.json()
                                if not commits:
                                    break # No more commits
                                
                                for c in commits:
                                    sha = c["sha"]
                                    if sha in seen_shas:
                                        continue # Skip duplicate
                                    seen_shas.add(sha)
                                    
                                    commit_date_str = c["commit"]["author"]["date"]
                                    dt_obj = datetime.strptime(commit_date_str, "%Y-%m-%dT%H:%M:%SZ")
                                    date_only = dt_obj.strftime("%Y-%m-%d")
                                    msg = c["commit"]["message"]
                                    
                                    all_commits.append({
                                        "date": date_only,
                                        "message": msg,
                                        "repo": repo
                                    })
                                
                                # Optimization: If fewer than 100 results, we reached end
                                if len(commits) < 100:
                                    break
                                page += 1
                                # Limit pages to prevent infinite loops on massive repos
                                if page > 20: 
                                    break
                                    
                            elif resp.status_code == 409:
                                break # Empty repo
                            else:
                                st.warning(f"Failed {repo}/{branch_label}: {resp.status_code}")
                                break
                            
                    except Exception as e:
                        st.error(f"Error fetching {repo}: {e}")
                
                progress_bar.progress((i + 1) / total_repos)

            status_text.text("Processing logs...")
            # Group by Date
            commits_by_date = {}
            for c in all_commits:
                d = c["date"]
                if d not in commits_by_date:
                    commits_by_date[d] = []
                commits_by_date[d].append(c)
            
            # Summarize
            if not commits_by_date:
                st.warning("No unique commits found matching your criteria.")
            else:
                generated_logs = []
                total_days = len(commits_by_date)
                
                gen_progress = st.progress(0, text="Summarizing with AI..." if groq_api_key else "Summarizing...")
                

                # Initialize Groq Client
                groq_client = None
                if groq_api_key:
                    try:
                        groq_client = Groq(api_key=groq_api_key)
                    except Exception as e:
                        st.error(f"Groq Init Error: {e}")

                def make_groq_request(client, prompt, retries=3):
                    """
                    Tries to get a completion with exponential backoff and model fallback.
                    Models: llama-3.1-8b-instant -> llama-3.3-70b-versatile
                    """
                    # Swapped order: 8b is faster and often has better rate limits
                    models = ["llama-3.1-8b-instant", "llama-3.3-70b-versatile"]
                    
                    for model in models:
                        for attempt in range(retries):
                            try:
                                return client.chat.completions.create(
                                    model=model,
                                    messages=[{"role": "user", "content": prompt}],
                                    temperature=0.3,
                                    max_tokens=6000, # More tokens for batch response
                                    response_format={"type": "json_object"} # STRICT JSON MODE
                                )
                            except Exception as e:
                                # Check for Rate Limit (429)
                                is_rate_limit = "429" in str(e) or (hasattr(e, 'status_code') and e.status_code == 429)
                                
                                if is_rate_limit:
                                    wait_time = 2 ** (attempt + 1) # 2, 4, 8 seconds
                                    if attempt < retries - 1:
                                        print(f"Rate limit on {model}. Retrying in {wait_time}s...")
                                        time.sleep(wait_time)
                                    else:
                                        print(f"Giving up on {model} after {retries} attempts.")
                                else:
                                    print(f"Error on {model}: {e}")
                                    break # Try next model immediately
                                    
                    return None # All failed

                # Prepare batches
                sorted_dates = sorted(commits_by_date.keys(), reverse=True)
                BATCH_SIZE = 3 
                batches = [sorted_dates[i:i + BATCH_SIZE] for i in range(0, len(sorted_dates), BATCH_SIZE)]
                
                total_batches = len(batches)
                
                for b_idx, batch_dates in enumerate(batches):
                    # Construct Prompt for the entire batch
                    batch_context = []
                    for d_str in batch_dates:
                        commits = commits_by_date[d_str]
                        msgs = [c["message"] for c in commits]
                        batch_context.append(f"Date: {d_str}\nCommits:\n" + "\n".join(f"- {m}" for m in msgs))
                    
                    full_batch_text = "\n\n".join(batch_context)
                    
                    if groq_client:
                        # Keep strict delay for safety, can reduce later if 8b proves robust
                        if b_idx > 0:
                            status_text.text(f"⏳ Throttling for 5s (Model Switch) to respect Rate Limits...")
                            time.sleep(5) # Reduced to 5s as 8b is generally lighter
                            status_text.text(f"Processing Batch {b_idx+1}/{total_batches}...")
                            
                        prompt = f"""Role: Professional software engineer writing daily logs for a university placement report.

Task:
For each date below, generate a SINGLE, concise log entry (max 20 words).
Output must be a valid JSON Object with a key "entries" containing the list.

Input Data:
{full_batch_text}

Output JSON Format:
{{
    "entries": [
        {{
            "date": "YYYY-MM-DD",
            "description": "Implemented login...",
            "activity_code": "4.2",
            "problem": "...",
            "solution": "..."
        }}
    ]
}}
"""
                        try:
                            response = make_groq_request(groq_client, prompt)
                            
                            if response:
                                text = response.choices[0].message.content
                                try:
                                    # With strict JSON mode, we should just load it directly
                                    data = json.loads(text)
                                    data_list = data.get("entries", [])
                                    
                                    # Map back to results
                                    for item in data_list:
                                        # Validate date exists in our batch
                                        if item.get("date") in batch_dates:
                                            generated_logs.append({
                                                "Date": item["date"],
                                                "Activity": item.get("activity_code", "4.2"),
                                                "Description": item.get("description", ""),
                                                "Problems": item.get("problem", ""),
                                                "Solutions": item.get("solution", "")
                                            })
                                except json.JSONDecodeError:
                                    st.warning(f"⚠️ JSON Parse Error for batch {b_idx+1}. Raw: {text[:100]}...")
                            else:
                                st.error(f"❌ Batch {b_idx+1} failed after retries.")
                                
                        except Exception as e:
                            st.warning(f"⚠️ Batch Error: {e}")

                    else:
                        # Fallback if no Groq Key (Manual fill)
                        for d_str in batch_dates:
                            commits = commits_by_date[d_str]
                            msgs = [c["message"] for c in commits]
                            repo_text = ", ".join(list(set([c["repo"] for c in commits])))
                            generated_logs.append({
                                "Date": d_str,
                                "Activity": "4.2",
                                "Description": f"Worked on {repo_text}. Commits: {msgs[0]}",
                                "Problems": "",
                                "Solutions": ""
                            })

                    # Update Progress
                    gen_progress.progress((b_idx + 1) / total_batches)

                generated_logs.sort(key=lambda x: x["Date"])
                st.session_state.generated_git_logs = pd.DataFrame(generated_logs)
                st.success(f"✅ Generated {len(generated_logs)} entries via GitHub API!")

    # 4. Preview & Save (Same as before)
    if "generated_git_logs" in st.session_state and not st.session_state.generated_git_logs.empty:
        st.subheader("Preview Generated Logs")
        edited_logs = st.data_editor(st.session_state.generated_git_logs, num_rows="dynamic")
        
        if st.button("💾 Save All Imported Logs"):
            count = 0
            for index, row in edited_logs.iterrows():
                save_entry(
                    datetime.strptime(row["Date"], "%Y-%m-%d"), 
                    "1.1", 
                    row["Description"], 
                    row["Problems"], 
                    row["Solutions"]
                )
                count += 1
            
            st.success(f"Successfully imported {count} logs!")
            st.session_state.generated_git_logs = pd.DataFrame()
            time.sleep(2)
            st.rerun()

# --- TAB 2: DAILY LOG ---
with tab_daily:
    st.header("📝 Daily Entry")
    
    # Initialize session state for daily form reset
    if "daily_form_key" not in st.session_state:
        st.session_state.daily_form_key = 0
    
    # Key suffix for resetting
    daily_key = str(st.session_state.daily_form_key)
    
    with st.form("daily_entry_form"):
        col1, col2 = st.columns(2)
        
        with col1:
            date_entry = st.date_input("Date", datetime.today(), key=f"date_{daily_key}")
        with col2:
            activity_code = st.selectbox(
                "Activity Code", 
                list(ACTIVITIES.keys()), 
                format_func=lambda x: f"{x} - {ACTIVITIES[x]}",
                key=f"act_code_{daily_key}"
            )
            
        description = st.text_area("Description of Work", height=100, key=f"daily_desc_{daily_key}")
        
        with st.expander("Problems & Solutions (Optional)"):
            prob = st.text_input("Problems Encountered", key=f"daily_prob_{daily_key}")
            sol = st.text_input("Solutions Finding", key=f"daily_sol_{daily_key}")
            
        submitted = st.form_submit_button("💾 Save Entry")
        
        if submitted:
            if description.strip():
                with st.spinner("Saving entry..."):
                    code = activity_code # activity_code is already the code, no split needed
                    save_entry(date_entry, code, description, prob, sol)
                    time.sleep(0.5) # Fake delay for UX
                
                st.success("✅ Entry Saved! Clearing form...")
                time.sleep(1)
                st.session_state.daily_form_key += 1
                st.rerun()
            else:
                st.error("⚠️ Description required!")


import time

# --- TAB 3: MANUAL BULK ENTRY ---
with tab_manual:
    st.header("📚 Bulk Week Entry")
    st.info("Select any day in a week. We'll load Monday to Friday for rapid entry.")
    
    # Initialize session state for form reset
    if "bulk_form_key" not in st.session_state:
        st.session_state.bulk_form_key = 0

    import calendar

    # Select Year and Month first
    col_y, col_m = st.columns(2)
    with col_y:
        current_year = datetime.today().year
        years = [current_year, current_year - 1]
        sel_year = st.selectbox("Year", years)
    
    with col_m:
        month_names = list(calendar.month_name)[1:]
        # Default to current month
        current_month_index = datetime.today().month - 1
        sel_month_name = st.selectbox("Month", month_names, index=current_month_index)
        sel_month = month_names.index(sel_month_name) + 1

    # Find all weeks (starting Monday) in the selected Month/Year
    week_options = []
    
    # Start checking from the 1st of the month
    # BUT we want to include a week even if it started in the previous month if the majority is in this month?
    # Or just strictly "Mondays in this month"?
    # Impl: "Mondays that fall within this month" is the clearest logic.
    
    d = datetime(sel_year, sel_month, 1)
    # Advance to the first Monday of the month
    while d.weekday() != 0: # 0 = Monday
        d += timedelta(days=1)
        
    # Collect all Mondays in this month
    while d.month == sel_month:
        w_end = d + timedelta(days=6)
        label = f"{d.strftime('%d %b')} - {w_end.strftime('%d %b %Y')}"
        week_options.append({"label": label, "start_date": d})
        d += timedelta(days=7)
    
    if not week_options:
        st.warning(f"No weeks start in {sel_month_name} {sel_year}.")
        start_of_week = datetime.today() # Fallback
    else:
        selected_option = st.selectbox(
            "Select Week", 
            week_options, 
            format_func=lambda x: x["label"]
        )
        start_of_week = selected_option["start_date"]
    
    end_of_week = start_of_week + timedelta(days=6)
    
    st.markdown(f"**Adding logs for: {start_of_week.strftime('%Y-%m-%d')} (Monday) to {end_of_week.strftime('%Y-%m-%d')} (Sunday)**")
    
    with st.form("bulk_entry_form"):
        entries = []
        activity_options = [f"{k} - {v}" for k, v in ACTIVITIES.items()]
        
        # KEY SUFFIX is crucial for resetting
        key_suffix = str(st.session_state.bulk_form_key)

        for i in range(7):
            day_date = start_of_week + timedelta(days=i)
            day_name = day_date.strftime("%A")
            
            st.markdown(f"---")
            st.subheader(f"{day_name} ({day_date.strftime('%d/%m')})")
            
            c1, c2 = st.columns([1, 2])
            with c1:
                act = st.selectbox(f"Activity ({day_name})", activity_options, key=f"act_{i}_{key_suffix}")
            with c2:
                desc = st.text_area(f"Description ({day_name})", height=70, key=f"desc_{i}_{key_suffix}", placeholder="Leave empty to skip")
            
            # Optional problems/solutions
            with st.expander(f"Problems & Solutions ({day_name})"):
                pc1, pc2 = st.columns(2)
                prob = pc1.text_input("Problem", key=f"prob_{i}_{key_suffix}")
                sol = pc2.text_input("Solution", key=f"sol_{i}_{key_suffix}")
            
            entries.append({
                "date": day_date,
                "activity": act,
                "description": desc,
                "problem": prob,
                "solution": sol
            })
        
        submitted = st.form_submit_button("💾 Save Full Week Logs")
        if submitted:
            with st.spinner("Saving entries..."):
                count = 0
                for entry in entries:
                    if entry["description"].strip():
                        code = entry["activity"].split(" - ")[0]
                        save_entry(
                            entry["date"], 
                            code, 
                            entry["description"], 
                            entry["problem"], 
                            entry["solution"]
                        )
                        count += 1
                
                time.sleep(0.5) # Fake delay for UX feel

            if count > 0:
                st.success(f"✅ Successfully saved {count} entries! Clearing form...")
                time.sleep(1)
                st.session_state.bulk_form_key += 1 # Increment key to reset widgets
                st.rerun()
            else:
                st.warning("⚠️ No descriptions entered. Nothing saved.")

# --- TAB 4: EXCEL AUTOMATOR ---
with tab_excel:
    st.header("Fill your IIT Record Book")
    st.info("The app will find empty weeks and fill them with your logs.")
    
    local_file_name = "Industrial Placement Record Book.xlsx"
    final_file = None
    
    # Check for local file
    if os.path.exists(local_file_name):
        st.success(f"📂 Found local file: **'{local_file_name}'**")
        if st.checkbox("Use this local file?", value=True):
            final_file = local_file_name
    
    # Fallback to uploader if no local file used
    if not final_file:
        final_file = st.file_uploader("Upload Record Book (.xlsx)", type=["xlsx"])
    
    col_d1, col_d2 = st.columns(2)
    with col_d1:
        gen_start_date = st.date_input("Generation Start Date", datetime(2025, 12, 15))
    with col_d2:
        gen_end_date = st.date_input("Generation End Date", datetime(2026, 1, 14))

    if final_file and not df.empty:
        if st.button("⚡ Fill Excel Sheet"):
            with st.spinner("Processing..."):
                # If using the local file directly, output to the same path
                save_path = local_file_name if (final_file == local_file_name) else None
                
                # Convert to datetime objects for function
                start_dt = datetime.combine(gen_start_date, datetime.min.time())
                end_dt = datetime.combine(gen_end_date, datetime.min.time())

                processed_excel, msg = fill_excel_sheet(final_file, df, start_dt, end_dt, output_path=save_path)
                
                if save_path and processed_excel is None:
                    # Direct save case
                    st.success(f"✅ Record Book updated directly! ({save_path})")
                    st.balloons()
                elif processed_excel:
                    # Buffer case (uploaded file)
                    st.success("Excel Filled Successfully!")
                    st.download_button(
                        label="📥 Download Updated Record Book",
                        data=processed_excel,
                        file_name="Updated_Record_Book.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                else:
                    st.error(msg)
    elif df.empty:
        st.warning("No logs found! Go to the 'Daily Log' tab and add some entries first.")

# --- TAB 5: HISTORY ---
with tab_hist:
    st.dataframe(df.sort_values(by="Date", ascending=False), use_container_width=True)
    if st.button("Clear All Data (Reset)"):
        if os.path.exists(FILE_NAME):
            os.remove(FILE_NAME)
            st.rerun()