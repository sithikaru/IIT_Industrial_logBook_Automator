import pandas as pd
import openpyxl
import os
import shutil
from app import fill_excel_sheet
from datetime import datetime, timedelta

# Setup paths
ORIGINAL_FILE = "Industrial Placement Record Book.xlsx"
TEST_TEMPLATE = "temp_template_dynamic.xlsx"
OUTPUT_FILE = "test_output_dynamic.xlsx"

# Ensure clean state
if os.path.exists(TEST_TEMPLATE):
    os.remove(TEST_TEMPLATE)
if os.path.exists(OUTPUT_FILE):
    os.remove(OUTPUT_FILE)

# Copy original to temp
shutil.copy(ORIGINAL_FILE, TEST_TEMPLATE)

# Create Mock Data
data = [{"Date": "2025-11-01", "Day": "SATURDAY", "Week_Ending": "2025-11-02", "Activity_Code": "1.1", "Description": "Test", "Problems": "", "Solutions": ""}]
df = pd.DataFrame(data)

# Test Range: Nov 2025 (has 5 Sundays: 2, 9, 16, 23, 30)
start_date = datetime(2025, 11, 1)
end_date = datetime(2025, 11, 30)

import traceback

print("Running fill_excel_sheet for Nov 2025 (5 Sundays)...")
try:
    fill_excel_sheet(TEST_TEMPLATE, df, start_date, end_date, output_path=OUTPUT_FILE)
    print("Function finished.")
except Exception as e:
    print(f"Error: {e}")
    traceback.print_exc()
    exit(1)

# Verification
print("Verifying output...")
wb = openpyxl.load_workbook(OUTPUT_FILE)

# Check Sheet Name
expected_sheet = "Nov 2025"
if expected_sheet not in wb.sheetnames:
    print(f"FAIL: Sheet '{expected_sheet}' not found. Sheets: {wb.sheetnames}")
    exit(1)

ws = wb[expected_sheet]

# Check for Week Endings count
found_weeks = []
for row in range(1, 200): # Scan enough rows
    val = ws.cell(row=row, column=2).value
    if val and isinstance(val, str) and "2025-11" in val:
        found_weeks.append(val)

print(f"Found weeks: {found_weeks}")

# Expectation: 5 Sundays
expected_count = 5

if len(found_weeks) != expected_count:
    print(f"FAIL: Expected {expected_count} weeks, found {len(found_weeks)}")
    exit(1)

print("SUCCESS: 5 Weeks generated correctly.")

# Cleanup
if os.path.exists(TEST_TEMPLATE):
    os.remove(TEST_TEMPLATE)
if os.path.exists(OUTPUT_FILE):
    os.remove(OUTPUT_FILE)
exit(0)
