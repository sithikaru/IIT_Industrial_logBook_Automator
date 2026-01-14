import openpyxl
import os

filename = 'Industrial Placement Record Book.xlsx'

if not os.path.exists(filename):
    print(f"Error: {filename} not found.")
    exit(1)

try:
    wb = openpyxl.load_workbook(filename)
    print(f"Sheet names: {wb.sheetnames}")

    if 'Logs' in wb.sheetnames:
        ws = wb['Logs']
        print("Using sheet: Logs")
    else:
        ws = wb.active
        print(f"Using active sheet: {ws.title}")

    print("-" * 30)
    print("Scanning for 'WEEK ENDING' anchors...")
    
    found = False
    for row in range(1, 100):
        c1 = ws.cell(row=row, column=1).value
        if c1 and "WEEK ENDING" in str(c1).upper():
            found = True
            c2 = ws.cell(row=row, column=2).value
            print(f"\n[FOUND] Row {row}: {c1} | {c2}")
            
            # Print next 12 rows to see the structure
            print("Structure below this anchor:")
            for i in range(1, 13):
                r_idx = row + i
                vals = [ws.cell(row=r_idx, column=c).value for c in range(1, 4)]
                print(f"  Row {r_idx}: {vals}")
                
    if not found:
        print("\nNo 'WEEK ENDING' found in the first 100 rows.")

except Exception as e:
    print(f"An error occurred: {e}")
